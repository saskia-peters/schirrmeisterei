from io import BytesIO

from fastapi import APIRouter, Depends, File, UploadFile
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.core.deps import get_admin_group_user, get_current_superuser, get_current_user
from app.core.exceptions import ConflictException, NotFoundException, ValidationException
from app.db.session import get_db
from app.models.models import (
    AppSetting,
    ConfigItem,
    ConfigItemType,
    EmailConfig,
    Organization,
    Permission,
    RolePermission,
    User,
    UserGroup,
)
from app.schemas.config import ConfigItemCreate, ConfigItemResponse, ConfigItemUpdate
from app.schemas.user import (
    AppSettingResponse,
    AppSettingUpdate,
    BulkUserUploadResult,
    EmailConfigCreate,
    EmailConfigResponse,
    EmailConfigUpdate,
    HierarchyUploadResult,
    PermissionResponse,
    RolePermissionUpdate,
    UserGroupAssignmentUpdate,
    UserGroupCreate,
    UserGroupDetailResponse,
    UserGroupResponse,
    UserGroupUpdate,
    UserResponse,
)
from app.services.organization_service import OrganizationService
from app.services.user_service import UserService

router = APIRouter(prefix="/admin", tags=["admin"])


@router.get("/config-items", response_model=list[ConfigItemResponse])
async def list_config_items(
    type: ConfigItemType | None = None,
    include_inactive: bool = False,
    db: AsyncSession = Depends(get_db),
    _: User = Depends(get_current_user),
) -> list[ConfigItem]:
    """Return all configuration items, optionally filtered by type and/or active status."""
    stmt = select(ConfigItem)
    if type is not None:
        stmt = stmt.where(ConfigItem.type == type)
    if not include_inactive:
        stmt = stmt.where(ConfigItem.is_active == True)  # noqa: E712
    stmt = stmt.order_by(ConfigItem.type, ConfigItem.sort_order, ConfigItem.name)
    result = await db.execute(stmt)
    return list(result.scalars().all())


@router.post("/config-items", response_model=ConfigItemResponse, status_code=201)
async def create_config_item(
    data: ConfigItemCreate,
    db: AsyncSession = Depends(get_db),
    _: User = Depends(get_current_superuser),
) -> ConfigItem:
    """Create a new configuration item."""
    item = ConfigItem(
        type=data.type,
        name=data.name,
        sort_order=data.sort_order,
    )
    db.add(item)
    await db.flush()
    await db.refresh(item)
    return item


@router.patch("/config-items/{item_id}", response_model=ConfigItemResponse)
async def update_config_item(
    item_id: str,
    data: ConfigItemUpdate,
    db: AsyncSession = Depends(get_db),
    _: User = Depends(get_current_superuser),
) -> ConfigItem:
    """Update an existing configuration item's name, sort order or active flag."""
    result = await db.execute(select(ConfigItem).where(ConfigItem.id == item_id))
    item = result.scalar_one_or_none()
    if item is None:
        raise NotFoundException("ConfigItem")
    if data.name is not None:
        item.name = data.name
    if data.sort_order is not None:
        item.sort_order = data.sort_order
    if data.is_active is not None:
        item.is_active = data.is_active
    await db.flush()
    await db.refresh(item)
    return item


@router.delete("/config-items/{item_id}", status_code=204)
async def delete_config_item(
    item_id: str,
    db: AsyncSession = Depends(get_db),
    _: User = Depends(get_current_superuser),
) -> None:
    """Delete a configuration item."""
    result = await db.execute(select(ConfigItem).where(ConfigItem.id == item_id))
    item = result.scalar_one_or_none()
    if item is None:
        raise NotFoundException("ConfigItem")
    await db.delete(item)
    await db.flush()


@router.get("/user-groups", response_model=list[UserGroupResponse])
async def list_user_groups(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_admin_group_user),
) -> list[UserGroupResponse]:
    """Return user groups for the current user's organization."""
    service = UserService(db)
    return await service.list_groups(current_user.organization_id)


@router.post("/user-groups", response_model=UserGroupResponse, status_code=201)
async def create_user_group(
    data: UserGroupCreate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_admin_group_user),
) -> UserGroupResponse:
    """Create a new user group within the current user's organization.

    Requires the manage_roles permission. Returns 409 if a group with that
    name already exists in the organization.
    """
    user_svc = UserService(db)
    if not current_user.is_superuser:
        has_perm = await user_svc.user_has_permission(current_user.id, "manage_roles")
        if not has_perm:
            raise ValidationException("You need the manage_roles permission")

    name = data.name.strip().lower()
    if not name:
        raise ValidationException("Group name cannot be empty")

    org_id = current_user.organization_id
    existing = await user_svc.get_group_by_name(name, org_id)
    if existing is not None:
        raise ConflictException("Group already exists")
    return await user_svc.create_group(name, org_id)


@router.patch("/user-groups/{group_id}", response_model=UserGroupResponse)
async def rename_user_group(
    group_id: str,
    data: UserGroupUpdate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_admin_group_user),
) -> UserGroupResponse:
    """Rename a user group. The helfende group cannot be renamed.

    Only groups belonging to the current user's organization can be renamed.
    """
    user_svc = UserService(db)
    if not current_user.is_superuser:
        has_perm = await user_svc.user_has_permission(current_user.id, "manage_roles")
        if not has_perm:
            raise ValidationException("You need the manage_roles permission")

    org_id = current_user.organization_id
    groups = await user_svc.list_groups(org_id)
    group = next((g for g in groups if g.id == group_id), None)
    if group is None:
        raise NotFoundException("UserGroup")

    name = data.name.strip().lower()
    if not name:
        raise ValidationException("Group name cannot be empty")
    existing = await user_svc.get_group_by_name(name, org_id)
    if existing is not None and existing.id != group.id:
        raise ConflictException("Group already exists")

    if group.name == "helfende" and name != "helfende":
        raise ValidationException("The helfende group cannot be renamed")

    group.name = name
    await db.flush()
    await db.refresh(group)
    return group


@router.delete("/user-groups/{group_id}", status_code=204)
async def delete_user_group(
    group_id: str,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_admin_group_user),
) -> None:
    """Delete a user group. Core groups (helfende, schirrmeister, admin) cannot be deleted.

    Only groups belonging to the current user's organization can be deleted.
    """
    user_svc = UserService(db)
    if not current_user.is_superuser:
        has_perm = await user_svc.user_has_permission(current_user.id, "manage_roles")
        if not has_perm:
            raise ValidationException("You need the manage_roles permission")

    org_id = current_user.organization_id
    groups = await user_svc.list_groups(org_id)
    group = next((g for g in groups if g.id == group_id), None)
    if group is None:
        raise NotFoundException("UserGroup")
    if group.name in {"helfende", "schirrmeister", "admin"}:
        raise ValidationException("Core groups cannot be deleted")
    await db.delete(group)
    await db.flush()


@router.get("/users/{user_id}/groups", response_model=list[str])
async def get_user_group_names(
    user_id: str,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_admin_group_user),
) -> list[str]:
    """Return the sorted list of group names the given user belongs to."""
    service = UserService(db)
    user = await service.get_by_id(user_id)
    if user is None:
        raise NotFoundException("User")
    # Non-superusers can only view users in their own org
    if not current_user.is_superuser and user.organization_id != current_user.organization_id:
        raise NotFoundException("User")
    return sorted([group.name for group in user.groups])


@router.put("/users/{user_id}/groups", response_model=list[str])
async def set_user_groups(
    user_id: str,
    data: UserGroupAssignmentUpdate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_admin_group_user),
) -> list[str]:
    """Replace the user's group memberships with the supplied set of group names.

    Groups are resolved within the user's organization scope.
    """
    service = UserService(db)
    user = await service.get_by_id(user_id)
    if user is None:
        raise NotFoundException("User")
    # Non-superusers can only manage users in their own org
    if not current_user.is_superuser and user.organization_id != current_user.organization_id:
        raise NotFoundException("User")
    try:
        updated = await service.assign_groups(user, set(data.group_names))
    except ValueError as exc:
        raise ValidationException(str(exc)) from exc
    return sorted([group.name for group in updated.groups])


@router.get("/users", response_model=list[UserResponse])
async def list_users_for_admin(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_admin_group_user),
) -> list[User]:
    """Return users visible to the current admin.

    Superusers see all users. Org admins see users in their organization.
    """
    service = UserService(db)
    if current_user.is_superuser:
        return await service.list_all()
    if current_user.organization_id:
        return await service.list_by_org(current_user.organization_id)
    return await service.list_all()


# ─── App Settings ──────────────────────────────────────────────────────────────

AGE_THRESHOLD_KEYS = [
    "age_green_days",
    "age_light_green_days",
    "age_yellow_days",
    "age_orange_days",
    "age_light_red_days",
]

AGE_THRESHOLD_DEFAULTS: dict[str, str] = {
    "age_green_days": "3",
    "age_light_green_days": "7",
    "age_yellow_days": "14",
    "age_orange_days": "21",
    "age_light_red_days": "30",
}


async def _ensure_age_defaults(db: AsyncSession) -> None:
    """Insert default age-threshold settings if they are not yet present in the database."""
    for key, default in AGE_THRESHOLD_DEFAULTS.items():
        result = await db.execute(select(AppSetting).where(AppSetting.key == key))
        if result.scalar_one_or_none() is None:
            db.add(AppSetting(key=key, value=default))
    await db.flush()


@router.get("/app-settings", response_model=list[AppSettingResponse])
async def list_app_settings(
    db: AsyncSession = Depends(get_db),
    _: User = Depends(get_current_user),
) -> list[AppSetting]:
    """Return all application settings, initialising age-threshold defaults if absent."""
    await _ensure_age_defaults(db)
    result = await db.execute(select(AppSetting).order_by(AppSetting.key))
    return list(result.scalars().all())


@router.patch("/app-settings/{key}", response_model=AppSettingResponse)
async def update_app_setting(
    key: str,
    data: AppSettingUpdate,
    db: AsyncSession = Depends(get_db),
    _: User = Depends(get_current_superuser),
) -> AppSetting:
    """Update the value of an application setting.

    Age-threshold values must be positive integers (number of days).
    """
    await _ensure_age_defaults(db)
    result = await db.execute(select(AppSetting).where(AppSetting.key == key))
    setting = result.scalar_one_or_none()
    if setting is None:
        raise NotFoundException("AppSetting")
    # For age threshold keys, validate that value is a positive integer
    if key in AGE_THRESHOLD_KEYS:
        try:
            days = int(data.value)
            if days < 1:
                raise ValueError
        except ValueError:
            raise ValidationException(
                "Age threshold must be a positive integer (number of days)"
            ) from None
    setting.value = data.value
    await db.flush()
    await db.refresh(setting)
    return setting


# ─── Permissions ───────────────────────────────────────────────────────────────

@router.get("/permissions", response_model=list[PermissionResponse])
async def list_permissions(
    db: AsyncSession = Depends(get_db),
    _: User = Depends(get_admin_group_user),
) -> list[Permission]:
    """Return all defined permissions ordered by codename."""
    result = await db.execute(select(Permission).order_by(Permission.codename))
    return list(result.scalars().all())


@router.get("/user-groups-detail", response_model=list[UserGroupDetailResponse])
async def list_user_groups_detail(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_admin_group_user),
) -> list[UserGroup]:
    """Return user groups with their associated permissions for the current user's org."""
    service = UserService(db)
    return await service.list_groups_detail(current_user.organization_id)


@router.put("/user-groups/{group_id}/permissions", response_model=UserGroupDetailResponse)
async def set_group_permissions(
    group_id: str,
    data: RolePermissionUpdate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_admin_group_user),
) -> UserGroup:
    """Replace all permissions for the given user group with the supplied set of codenames.

    Requires the manage_roles permission. Only groups within the current user's
    organization can be modified.
    """
    user_svc = UserService(db)
    if not current_user.is_superuser:
        has_perm = await user_svc.user_has_permission(current_user.id, "manage_roles")
        if not has_perm:
            raise ValidationException("You need the manage_roles permission")

    # Verify group belongs to current user's org
    org_id = current_user.organization_id
    org_groups = await user_svc.list_groups(org_id)
    if not any(g.id == group_id for g in org_groups):
        raise NotFoundException("UserGroup")

    result = await db.execute(
        select(UserGroup).where(UserGroup.id == group_id)
    )
    group = result.scalar_one_or_none()
    if group is None:
        raise NotFoundException("UserGroup")

    # Delete existing role-permission links
    existing = await db.execute(
        select(RolePermission).where(RolePermission.role_id == group_id)
    )
    for rp in existing.scalars().all():
        await db.delete(rp)
    await db.flush()

    # Add new permissions
    if data.permission_codenames:
        perm_result = await db.execute(
            select(Permission).where(Permission.codename.in_(data.permission_codenames))
        )
        permissions = list(perm_result.scalars().all())
        for perm in permissions:
            db.add(RolePermission(role_id=group_id, permission_id=perm.id))
        await db.flush()

    # Reload with permissions
    result = await db.execute(
        select(UserGroup)
        .options(selectinload(UserGroup.permissions))
        .where(UserGroup.id == group_id)
    )
    return result.scalar_one()


# ─── Email Config ──────────────────────────────────────────────────────────────

@router.get("/email-configs", response_model=list[EmailConfigResponse])
async def list_email_configs(
    db: AsyncSession = Depends(get_db),
    _: User = Depends(get_admin_group_user),
) -> list[EmailConfig]:
    """Return all email configurations with their associated organisations."""
    result = await db.execute(
        select(EmailConfig)
        .options(selectinload(EmailConfig.organization))
        .order_by(EmailConfig.organization_id)
    )
    return list(result.scalars().all())


@router.get("/email-configs/{org_id}", response_model=EmailConfigResponse)
async def get_email_config(
    org_id: str,
    db: AsyncSession = Depends(get_db),
    _: User = Depends(get_admin_group_user),
) -> EmailConfig:
    """Fetch the email configuration for the given organisation, or 404 if not found."""
    result = await db.execute(
        select(EmailConfig)
        .options(selectinload(EmailConfig.organization))
        .where(EmailConfig.organization_id == org_id)
    )
    config = result.scalar_one_or_none()
    if config is None:
        raise NotFoundException("EmailConfig")
    return config


@router.post("/email-configs", response_model=EmailConfigResponse, status_code=201)
async def create_email_config(
    data: EmailConfigCreate,
    db: AsyncSession = Depends(get_db),
    _: User = Depends(get_admin_group_user),
) -> EmailConfig:
    """Create an email configuration for an organisation. Returns 409 if one already exists."""
    existing = await db.execute(
        select(EmailConfig).where(EmailConfig.organization_id == data.organization_id)
    )
    if existing.scalar_one_or_none() is not None:
        raise ConflictException("Email config already exists for this organization")
    config = EmailConfig(
        organization_id=data.organization_id,
        smtp_host=data.smtp_host,
        smtp_port=data.smtp_port,
        smtp_user=data.smtp_user,
        smtp_password=data.smtp_password,
        from_email=data.from_email,
        use_tls=data.use_tls,
        is_active=data.is_active,
    )
    db.add(config)
    await db.flush()
    await db.refresh(config, attribute_names=["organization"])
    return config


@router.patch("/email-configs/{config_id}", response_model=EmailConfigResponse)
async def update_email_config(
    config_id: str,
    data: EmailConfigUpdate,
    db: AsyncSession = Depends(get_db),
    _: User = Depends(get_admin_group_user),
) -> EmailConfig:
    """Update an existing email configuration's SMTP settings."""
    result = await db.execute(select(EmailConfig).where(EmailConfig.id == config_id))
    config = result.scalar_one_or_none()
    if config is None:
        raise NotFoundException("EmailConfig")
    if data.smtp_host is not None:
        config.smtp_host = data.smtp_host
    if data.smtp_port is not None:
        config.smtp_port = data.smtp_port
    if data.smtp_user is not None:
        config.smtp_user = data.smtp_user
    if data.smtp_password is not None:
        config.smtp_password = data.smtp_password
    if data.from_email is not None:
        config.from_email = data.from_email
    if data.use_tls is not None:
        config.use_tls = data.use_tls
    if data.is_active is not None:
        config.is_active = data.is_active
    await db.flush()
    await db.refresh(config, attribute_names=["organization"])
    return config


# ─── Bulk User Upload ─────────────────────────────────────────────────────────

@router.post("/users/bulk-upload", response_model=BulkUserUploadResult)
async def bulk_upload_users(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_admin_group_user),
) -> BulkUserUploadResult:
    """Upload an XLSX file to create multiple users at once.
    Required columns: email, full_name, password
    Optional columns: organization_id
    """
    if not file.filename or not file.filename.endswith(".xlsx"):
        raise ValidationException("File must be an XLSX file")

    from openpyxl import load_workbook

    contents = await file.read()
    wb = load_workbook(BytesIO(contents), read_only=True)
    ws = wb.active
    if ws is None:
        raise ValidationException("Empty workbook")

    # Read header row
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        raise ValidationException("Empty spreadsheet")

    header = [str(c).strip().lower() if c else "" for c in rows[0]]
    required = {"email", "full_name", "password"}
    if not required.issubset(set(header)):
        raise ValidationException(f"Missing required columns: {required - set(header)}")

    email_idx = header.index("email")
    name_idx = header.index("full_name")
    pwd_idx = header.index("password")
    org_idx = header.index("organization_id") if "organization_id" in header else None

    service = UserService(db)
    created = 0
    errors: list[str] = []

    for i, row in enumerate(rows[1:], start=2):
        email = str(row[email_idx]).strip() if row[email_idx] else ""
        full_name = str(row[name_idx]).strip() if row[name_idx] else ""
        password = str(row[pwd_idx]).strip() if row[pwd_idx] else ""

        if not email or not full_name or not password:
            errors.append(f"Row {i}: missing required fields")
            continue

        if len(password) < 8:
            errors.append(f"Row {i}: password must be at least 8 characters")
            continue

        existing = await service.get_by_email(email)
        if existing:
            errors.append(f"Row {i}: email {email} already exists")
            continue

        org_id = None
        if org_idx is not None and row[org_idx]:
            org_id = str(row[org_idx]).strip()
        elif current_user.organization_id:
            org_id = current_user.organization_id

        from app.schemas.user import UserCreate
        user_data = UserCreate(
            email=email,
            full_name=full_name,
            password=password,
            organization_id=org_id,
        )
        try:
            await service.create(user_data)
            created += 1
        except Exception as exc:
            errors.append(f"Row {i}: {exc}")

    return BulkUserUploadResult(created=created, errors=errors)


# ─── Hierarchy Upload ─────────────────────────────────────────────────────────

@router.post("/hierarchy/upload", response_model=HierarchyUploadResult)
async def upload_hierarchy(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    _: User = Depends(get_current_superuser),
) -> HierarchyUploadResult:
    """Upload an XLSX file to import the organisation hierarchy.
    Required columns: level, name, parent_name
    level must be one of: leitung, landesverband, regionalstelle, ortsverband
    parent_name is the display name of the parent unit (empty for top-level).
    Existing orgs with the same name+level are skipped (not updated).
    """
    from app.models.models import OrganizationLevel

    if not file.filename or not file.filename.endswith(".xlsx"):
        raise ValidationException("File must be an XLSX file")

    from openpyxl import load_workbook

    contents = await file.read()
    wb = load_workbook(BytesIO(contents), read_only=True)
    ws = wb.active
    if ws is None:
        raise ValidationException("Empty workbook")

    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        raise ValidationException("Empty spreadsheet")

    header = [str(c).strip().lower() if c else "" for c in rows[0]]
    required = {"level", "name", "parent_name"}
    if not required.issubset(set(header)):
        raise ValidationException(f"Missing required columns: {required - set(header)}")

    level_idx = header.index("level")
    name_idx = header.index("name")
    parent_idx = header.index("parent_name")

    LEVEL_ORDER = ["leitung", "landesverband", "regionalstelle", "ortsverband"]
    data_rows: list[tuple[str, str, str]] = []
    for row in rows[1:]:
        lvl = str(row[level_idx]).strip().lower() if row[level_idx] else ""
        nm = str(row[name_idx]).strip() if row[name_idx] else ""
        parent_nm = str(row[parent_idx]).strip() if row[parent_idx] else ""
        if lvl and nm:
            data_rows.append((lvl, nm, parent_nm))

    # Sort by level order so parents are created first
    data_rows.sort(key=lambda r: LEVEL_ORDER.index(r[0]) if r[0] in LEVEL_ORDER else 99)

    service = OrganizationService(db)
    created = 0
    skipped = 0
    errors: list[str] = []

    for lvl, nm, parent_nm in data_rows:
        try:
            org_level = OrganizationLevel(lvl)
        except ValueError:
            errors.append(f"Unknown level '{lvl}' for org '{nm}'")
            continue

        existing = await service.find_by_name_level(nm, org_level)
        if existing is not None:
            skipped += 1
            continue

        parent_id: str | None = None
        if parent_nm:
            # Search for parent across all levels above current
            parent_org = await db.execute(
                select(Organization).where(Organization.name == parent_nm)
            )
            parent = parent_org.scalar_one_or_none()
            if parent is None:
                errors.append(f"Parent '{parent_nm}' not found for '{nm}' — skipped")
                continue
            parent_id = parent.id

        await service.create_org(org_level, nm, parent_id)
        created += 1

    await db.commit()
    return HierarchyUploadResult(created=created, skipped=skipped, errors=errors)
