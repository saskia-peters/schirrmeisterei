"""multi-tenancy: organizations, permissions, email config

Revision ID: 0002
Revises: 0001
Create Date: 2026-04-05

"""
import uuid
from collections.abc import Sequence
from datetime import datetime, timezone
from pathlib import Path

import bcrypt
import sqlalchemy as sa
from alembic import op

revision: str = "0002"
down_revision: str | None = "0001"
branch_labels: str | Sequence[str] | None = None
depends_on: str | Sequence[str] | None = None

# ── Permission definitions ────────────────────────────────────────────────────

PERMISSIONS = [
    ("create_ticket", "Create new tickets"),
    ("view_ticket", "View tickets"),
    ("edit_ticket", "Edit ticket details"),
    ("close_ticket", "Close tickets"),
    ("delete_ticket", "Delete tickets"),
    ("assign_ticket", "Assign tickets to users"),
    ("add_comment", "Add comments to tickets"),
    ("delete_comment", "Delete comments from tickets"),
    ("upload_attachment", "Upload attachments to tickets"),
    ("delete_attachment", "Delete attachments from tickets"),
    ("manage_users", "Manage users in the organization"),
    ("manage_config", "Manage system configuration"),
    ("manage_email", "Manage email configuration"),
    ("bulk_upload_users", "Upload users via XLSX"),
]

ROLE_PERMISSIONS: dict[str, list[str]] = {
    "helfende": [
        "create_ticket", "view_ticket", "edit_ticket", "add_comment", "upload_attachment",
    ],
    "schirrmeister": [
        "create_ticket", "view_ticket", "edit_ticket", "close_ticket",
        "assign_ticket", "add_comment", "delete_comment",
        "upload_attachment", "delete_attachment",
    ],
    "admin": [
        "create_ticket", "view_ticket", "edit_ticket", "close_ticket",
        "assign_ticket", "add_comment", "delete_comment",
        "upload_attachment", "delete_attachment",
        "manage_users", "manage_config", "manage_email",
        "delete_ticket", "bulk_upload_users",
    ],
}


def _load_hierarchy() -> list[tuple[str, str, str]]:
    """Read the XLSX hierarchy file. Returns list of (level, name, parent_name)."""
    from openpyxl import load_workbook

    xlsx = Path(__file__).resolve().parents[2] / "data" / "organisation_hierarchy.xlsx"
    wb = load_workbook(xlsx, read_only=True)
    ws = wb.active
    assert ws is not None
    rows: list[tuple[str, str, str]] = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue  # skip header
        level, name, parent_name = str(row[0]), str(row[1]), str(row[2] or "")
        rows.append((level, name, parent_name))
    wb.close()
    return rows


def upgrade() -> None:
    now = datetime.now(timezone.utc)

    # ── organizations ─────────────────────────────────────────────────────────
    op.create_table(
        "organizations",
        sa.Column("id", sa.String(36), primary_key=True),
        sa.Column("name", sa.String(255), nullable=False),
        sa.Column(
            "level",
            sa.Enum("ortsverband", "regionalstelle", "landesverband", "leitung",
                    name="organizationlevel"),
            nullable=False,
        ),
        sa.Column("parent_id", sa.String(36), sa.ForeignKey("organizations.id"), nullable=True),
        sa.Column("created_at", sa.DateTime(timezone=True), nullable=False),
    )
    op.create_index("ix_organizations_level", "organizations", ["level"])
    op.create_index("ix_organizations_parent_id", "organizations", ["parent_id"])

    # ── permissions ───────────────────────────────────────────────────────────
    op.create_table(
        "permissions",
        sa.Column("id", sa.String(36), primary_key=True),
        sa.Column("codename", sa.String(100), nullable=False, unique=True),
        sa.Column("description", sa.String(255), nullable=False),
        sa.Column("created_at", sa.DateTime(timezone=True), nullable=False),
    )
    op.create_index("ix_permissions_codename", "permissions", ["codename"], unique=True)

    # ── role_permissions ──────────────────────────────────────────────────────
    op.create_table(
        "role_permissions",
        sa.Column("role_id", sa.String(36), sa.ForeignKey("user_groups.id"), primary_key=True),
        sa.Column("permission_id", sa.String(36), sa.ForeignKey("permissions.id"), primary_key=True),
        sa.Column("created_at", sa.DateTime(timezone=True), nullable=False),
    )

    # ── email_configs ─────────────────────────────────────────────────────────
    op.create_table(
        "email_configs",
        sa.Column("id", sa.String(36), primary_key=True),
        sa.Column("organization_id", sa.String(36), sa.ForeignKey("organizations.id"),
                  nullable=False, unique=True),
        sa.Column("smtp_host", sa.String(255), nullable=False, server_default=""),
        sa.Column("smtp_port", sa.Integer, nullable=False, server_default="587"),
        sa.Column("smtp_user", sa.String(255), nullable=False, server_default=""),
        sa.Column("smtp_password", sa.String(255), nullable=False, server_default=""),
        sa.Column("from_email", sa.String(255), nullable=False, server_default=""),
        sa.Column("use_tls", sa.Boolean, nullable=False, server_default="1"),
        sa.Column("is_active", sa.Boolean, nullable=False, server_default="0"),
        sa.Column("created_at", sa.DateTime(timezone=True), nullable=False),
        sa.Column("updated_at", sa.DateTime(timezone=True), nullable=False),
    )

    # ── add organization_id to users ──────────────────────────────────────────
    op.add_column("users", sa.Column("organization_id", sa.String(36), nullable=True))
    op.create_foreign_key(
        "fk_users_organization_id", "users", "organizations",
        ["organization_id"], ["id"]
    )
    op.create_index("ix_users_organization_id", "users", ["organization_id"])

    # ── add organization_id to tickets ────────────────────────────────────────
    op.add_column("tickets", sa.Column("organization_id", sa.String(36), nullable=True))
    op.create_foreign_key(
        "fk_tickets_organization_id", "tickets", "organizations",
        ["organization_id"], ["id"]
    )
    op.create_index("ix_tickets_organization_id", "tickets", ["organization_id"])

    # ══════════════════════════════════════════════════════════════════════════
    # SEED DATA
    # ══════════════════════════════════════════════════════════════════════════

    # ── seed: organizations from XLSX ─────────────────────────────────────────
    hierarchy = _load_hierarchy()
    org_ids: dict[str, str] = {}  # name -> id

    for level, name, parent_name in hierarchy:
        oid = str(uuid.uuid4())
        org_ids[name] = oid
        conn = op.get_bind()
        conn.execute(
            sa.text(
                "INSERT INTO organizations (id, name, level, parent_id, created_at) "
                "VALUES (:id, :name, CAST(:level AS organizationlevel), :parent_id, :created_at)"
            ),
            {
                "id": oid,
                "name": name,
                "level": level,
                "parent_id": org_ids.get(parent_name),
                "created_at": now,
            },
        )

    # ── seed: permissions ─────────────────────────────────────────────────────
    perm_ids: dict[str, str] = {}
    perms_table = sa.table(
        "permissions",
        sa.column("id", sa.String),
        sa.column("codename", sa.String),
        sa.column("description", sa.String),
        sa.column("created_at", sa.DateTime(timezone=True)),
    )
    for codename, description in PERMISSIONS:
        pid = str(uuid.uuid4())
        perm_ids[codename] = pid
        op.bulk_insert(
            perms_table,
            [{"id": pid, "codename": codename, "description": description, "created_at": now}],
        )

    # ── seed: role_permissions ────────────────────────────────────────────────
    # Look up existing role (user_group) IDs
    conn = op.get_bind()
    role_rows = conn.execute(sa.text("SELECT id, name FROM user_groups")).fetchall()
    role_ids: dict[str, str] = {row[1]: row[0] for row in role_rows}

    rp_table = sa.table(
        "role_permissions",
        sa.column("role_id", sa.String),
        sa.column("permission_id", sa.String),
        sa.column("created_at", sa.DateTime(timezone=True)),
    )
    for role_name, perm_codenames in ROLE_PERMISSIONS.items():
        rid = role_ids.get(role_name)
        if not rid:
            continue
        for codename in perm_codenames:
            pid = perm_ids.get(codename)
            if not pid:
                continue
            op.bulk_insert(
                rp_table,
                [{"role_id": rid, "permission_id": pid, "created_at": now}],
            )

    # ── seed: assign existing admin user to the Leitung org ───────────────────
    leitung_id = org_ids.get("THW-Leitung")
    if leitung_id:
        conn.execute(
            sa.text("UPDATE users SET organization_id = :org_id WHERE is_superuser = true"),
            {"org_id": leitung_id},
        )

    # ── seed: second superadmin user ─────────────────────────────────────────
    superadmin2_id = str(uuid.uuid4())
    hashed2 = bcrypt.hashpw(b"superadmin", bcrypt.gensalt()).decode()

    users_t = sa.table(
        "users",
        sa.column("id", sa.String),
        sa.column("email", sa.String),
        sa.column("hashed_password", sa.String),
        sa.column("full_name", sa.String),
        sa.column("is_active", sa.Boolean),
        sa.column("is_superuser", sa.Boolean),
        sa.column("force_password_change", sa.Boolean),
        sa.column("totp_secret", sa.String),
        sa.column("totp_enabled", sa.Boolean),
        sa.column("organization_id", sa.String),
        sa.column("created_at", sa.DateTime(timezone=True)),
        sa.column("updated_at", sa.DateTime(timezone=True)),
    )
    op.bulk_insert(
        users_t,
        [{
            "id": superadmin2_id,
            "email": "superadmin@example.com",
            "hashed_password": hashed2,
            "full_name": "Super Admin",
            "is_active": True,
            "is_superuser": True,
            "force_password_change": True,
            "totp_secret": None,
            "totp_enabled": False,
            "organization_id": leitung_id,
            "created_at": now,
            "updated_at": now,
        }],
    )

    # Assign second superadmin to helfende + admin groups
    helfende_id = role_ids.get("helfende")
    admin_gid = role_ids.get("admin")
    memberships_t = sa.table(
        "user_group_memberships",
        sa.column("user_id", sa.String),
        sa.column("group_id", sa.String),
        sa.column("created_at", sa.DateTime(timezone=True)),
    )
    rows_to_add = []
    if helfende_id:
        rows_to_add.append({"user_id": superadmin2_id, "group_id": helfende_id, "created_at": now})
    if admin_gid:
        rows_to_add.append({"user_id": superadmin2_id, "group_id": admin_gid, "created_at": now})
    if rows_to_add:
        op.bulk_insert(memberships_t, rows_to_add)

    # ── Make tickets.organization_id NOT NULL after back-fill ─────────────────
    # (Any existing tickets without an org will fail.  In a fresh DB there are
    #  none, but we set it nullable=True above for safety during migration.)
    # We leave it nullable for now to accommodate existing data.


def downgrade() -> None:
    op.drop_index("ix_tickets_organization_id", table_name="tickets")
    op.drop_constraint("fk_tickets_organization_id", "tickets", type_="foreignkey")
    op.drop_column("tickets", "organization_id")

    op.drop_index("ix_users_organization_id", table_name="users")
    op.drop_constraint("fk_users_organization_id", "users", type_="foreignkey")
    op.drop_column("users", "organization_id")

    op.drop_table("email_configs")
    op.drop_table("role_permissions")
    op.drop_index("ix_permissions_codename", table_name="permissions")
    op.drop_table("permissions")
    op.drop_index("ix_organizations_parent_id", table_name="organizations")
    op.drop_index("ix_organizations_level", table_name="organizations")
    op.drop_table("organizations")
